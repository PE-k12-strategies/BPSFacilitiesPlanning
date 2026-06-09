"""
Build data/processed/ese_feeder_matrix.json from the K-12 feeder plan workbook.
Matrix row = school MSID (column A); program columns = destinations for students from that school.
Reverse index 'acceptsFrom': schools whose feeder cell lists this MSID as a destination.
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "data" / "sourcedocs" / "k12_feeder_plan_2026_2027_revised_2026_04_07.xlsx"
OUT_JSON = ROOT / "data" / "processed" / "ese_feeder_matrix.json"

PROGRAM_KEYS = (
    "ve_s",
    "ve_p",
    "ve_b",
    "day_school",
    "dhh",
    "vi",
    "blast",
)


def parse_msids(raw: object) -> list[str]:
    if raw is None:
        return []
    t = str(raw).strip()
    if not t or t.upper() == "N/A":
        return []
    found: list[str] = []
    for m in re.finditer(r"\d{2,5}", t):
        found.append(str(int(m.group(0))))
    if re.search(r"(?<!\d)11(?!\d)", t):
        if "11" not in found:
            found.append("11")
    seen: set[str] = set()
    out: list[str] = []
    for x in found:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 10) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip().lower() == "school":
            return r
    raise ValueError("Could not find 'School' header cell in column A")


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT_JSON

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = find_header_row(ws)

    headers_full: dict[str, str] = {}
    for i, key in enumerate(PROGRAM_KEYS):
        col = 2 + i
        h = ws.cell(hr, col).value
        headers_full[key] = (str(h).strip() if h is not None else "").replace("\r\n", "\n")

    rows_out: dict[str, dict[str, list[str]]] = {}
    accepts: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    for r in range(hr + 1, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None or str(sid).strip() == "":
            continue
        try:
            school_msid = str(int(float(str(sid))))
        except (TypeError, ValueError):
            continue
        prog_map: dict[str, list[str]] = {}
        for i, key in enumerate(PROGRAM_KEYS):
            col = 2 + i
            dests = parse_msids(ws.cell(r, col).value)
            prog_map[key] = dests
            for d in dests:
                accepts[d][key].add(school_msid)
        rows_out[school_msid] = prog_map

    accepts_out: dict[str, dict[str, list[str]]] = {}
    for target, by_prog in accepts.items():
        accepts_out[target] = {
            k: sorted(by_prog.get(k, []), key=lambda x: int(x)) for k in PROGRAM_KEYS
        }

    programs_meta = [
        {"key": "ve_s", "shortLabel": "VE-S", "headerFull": headers_full["ve_s"]},
        {"key": "ve_p", "shortLabel": "VE-P", "headerFull": headers_full["ve_p"]},
        {"key": "ve_b", "shortLabel": "VE-B", "headerFull": headers_full["ve_b"]},
        {"key": "day_school", "shortLabel": "Day School", "headerFull": headers_full["day_school"]},
        {"key": "dhh", "shortLabel": "Deaf or Hard of Hearing", "headerFull": headers_full["dhh"]},
        {"key": "vi", "shortLabel": "Visually Impaired", "headerFull": headers_full["vi"]},
        {"key": "blast", "shortLabel": "BLAST", "headerFull": headers_full["blast"]},
    ]

    payload = {
        "source": xlsx_path.name,
        "programs": programs_meta,
        "rows": rows_out,
        "acceptsFrom": accepts_out,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print("Wrote", out_path, "schools", len(rows_out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
