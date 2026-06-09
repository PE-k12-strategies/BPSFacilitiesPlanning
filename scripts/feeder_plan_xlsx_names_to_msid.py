"""
Apply school-name → MSID replacements to a feeder-plan Excel workbook (same rules as CSV script).
Writes back to the same path (or optional output path).
"""

from __future__ import annotations

import math
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import feeder_plan_csv_names_to_msid as fm


def coerce_cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return str(value).rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value)


def find_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 1).value
        if c is not None and str(c).strip().lower() == "school":
            return r
    raise ValueError("No row with first cell 'School' found")


def main() -> int:
    default = (
        Path(__file__).resolve().parent.parent
        / "data"
        / "sourcedocs"
        / "k12_feeder_plan_2026_2027_revised_2026_04_07.xlsx"
    )
    inp = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    outp = Path(sys.argv[2]) if len(sys.argv) > 2 else inp

    pairs = fm.build_replacement_pairs(fm.load_master_rows())

    wb = load_workbook(inp)
    ws = wb[wb.sheetnames[0]]
    hr = find_header_row(ws)

    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row):
        for cell in row:
            raw = coerce_cell_to_str(cell.value)
            if fm.should_skip_cell(raw):
                cell.value = raw if raw else None
                continue
            out = fm.replace_cell(raw, pairs)
            cell.value = out if out else None

    wb.save(outp)
    print("Wrote", outp.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
