# Builds data/processed/travel_impact.json from the three Travel Impact workbooks.
# Rows: compact triples [attendance_msid, scenario_msid, total_length_ft].
# Run from repo root: py -3 scripts/export_travel_impact_from_xlsx.py
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

# Add (scenario_destination_msid, "WorkbookName.xlsx") pairs as workbooks arrive in data/raw/.
MAPPING = (
    (3031, "JohnsonTravelImpact.xlsx"),
    (1081, "McNairTravelImpact.xlsx"),
    (2071, "StoneTravelImpact.xlsx"),
)


def _norm_msid(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v != v:  # NaN
            return None
        i = int(v)
        return i if i == v or abs(v - i) < 1e-6 else None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _norm_ft(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v != v:
            return None
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_workbook(path: Path) -> list[list[int | float]]:
    """Returns rows as [attendance, scenario, ft]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        wb.close()
        return []

    cols: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            cols[name.lower()] = i

    def col(*name_variants: str) -> int | None:
        for nm in name_variants:
            j = cols.get(nm.lower())
            if j is not None:
                return j
        return None

    ci_att = col("attendance msid", "attendance_msid")
    ci_sce = col("scenario msid", "scenario_msid")
    ci_ft = col("total_length")

    if ci_att is None or ci_sce is None or ci_ft is None:
        # Fallback: B, C, G (1-based) => 0-based 1,2,6
        ci_att, ci_sce, ci_ft = 1, 2, 6

    out: list[list[int | float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        a = _norm_msid(row[ci_att] if ci_att < len(row) else None)
        s = _norm_msid(row[ci_sce] if ci_sce < len(row) else None)
        ft = _norm_ft(row[ci_ft] if ci_ft < len(row) else None)
        if a is None or s is None or ft is None:
            continue
        out.append([a, s, round(ft, 6)])
    wb.close()
    return out


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    raw_dir = root / "data" / "raw"
    out_path = root / "data" / "processed" / "travel_impact.json"

    bundle: dict = {"schemaVersion": "1.0", "byMsid": {}}

    for key_msid, fname in MAPPING:
        path = raw_dir / fname
        if not path.is_file():
            print(f"Skipping (not found): {path}", file=sys.stderr)
            continue
        rows = read_workbook(path)
        bundle["byMsid"][str(key_msid)] = {
            "sourceFile": fname,
            "rowCount": len(rows),
            "rows": rows,
        }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(bundle, f, separators=(",", ":"))

    print(f"Wrote {out_path}")
    for k, info in bundle["byMsid"].items():
        print(f"  MSID {k}: {info['rowCount']} rows from {info['sourceFile']}")


if __name__ == "__main__":
    main()
