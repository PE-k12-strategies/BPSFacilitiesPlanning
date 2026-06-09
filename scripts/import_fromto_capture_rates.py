# From-To workbooks -> assignment / other district / choice / charter capture decimals in data/school_master.csv
#
# Layout (251010 workbooks): first sheet; row 6 = school type band; row 7 = enrollment column headers;
# residence labels are in column C (header "Students' School of Residence" on row 7);
# some rows leave C blank and repeat the label in column D (e.g. Viera Elem / Viera Middle).
# Denominator column = row 7 text containing "Residents in Boundary" (BK≈63 for Elem; W≈23 for Middle; V≈22 for High).
#
# Usage:
#   py -3 scripts/import_fromto_capture_rates.py
#   py -3 scripts/import_fromto_capture_rates.py --elem P:\\...\\251010FromToElem.xlsx ...
#
# Requires openpyxl. Run on a machine with the P: workbooks available.

from __future__ import annotations

import argparse
import csv
import importlib.util
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("Requires openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "data" / "school_master.csv"
REPORT_PATH = ROOT / "data" / "processed" / "fromto_capture_import_report.txt"

DEFAULT_ELEM = r"P:\0109260\Planning\WorkingFiles\00_To Be Sorted\FromToAnalysis\251010FromToElem.xlsx"
DEFAULT_MID = r"P:\0109260\Planning\WorkingFiles\00_To Be Sorted\FromToAnalysis\251010FromToMiddle.xlsx"
DEFAULT_HIGH = r"P:\0109260\Planning\WorkingFiles\00_To Be Sorted\FromToAnalysis\251010FromToHigh.xlsx"


def load_charter_matchers():
    spec = importlib.util.spec_from_file_location(
        "icc", ROOT / "scripts" / "import_charter_capture_from_xlsx.py"
    )
    icc = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(icc)
    return icc


icc = load_charter_matchers()


def canon_residence(s: str) -> str:
    s = str(s).replace("\r", " ").replace("\n", " ")
    return " ".join(s.split()).strip().casefold()


# find_match fails for these HS labels (ambiguous pair match); MS file has blank gap before jr/sr rows.
def _z4(msid: str) -> str:
    return str(int(str(msid).strip())).zfill(4)


MANUAL_RESIDENCE_TO_MSID: dict[str, str] = {
    canon_residence("Cocoa Jr/Sr (7-8 Only)"): _z4("1121"),
    canon_residence("Cocoa Jr/Sr (9-12 Only)"): _z4("1121"),
    canon_residence("Cocoa Beach (7-8 Only)"): _z4("5011"),
    canon_residence("Cocoa Beach (9-12 Only)"): _z4("5011"),
    canon_residence("Space Coast (7-8 Only)"): _z4("302"),
    canon_residence("Space Coast (9-12 Only)"): _z4("302"),
    # From-To workbooks use short labels not matched by find_match()
    canon_residence("Viera Elem"): _z4("3161"),
    canon_residence("Viera Middle"): _z4("3171"),
}


def pool_for_tier(schools: list[dict], tier: str) -> list[dict]:
    p = icc.pool_by_tier(schools)
    if tier == "elementary":
        return p["elementary"]
    if tier == "middle":
        return p["middle"] + [
            s for s in schools if (s.get("school_level") or "").lower() == "jr_sr_high"
        ]
    if tier == "high":
        return p["high"] + [
            s for s in schools if (s.get("school_level") or "").lower() == "jr_sr_high"
        ]
    return []


def match_residence_to_msid(
    residence: str, tier: str, schools: list[dict]
) -> tuple[str | None, str | None]:
    key = canon_residence(residence)
    if key in MANUAL_RESIDENCE_TO_MSID:
        return MANUAL_RESIDENCE_TO_MSID[key], None
    pool = pool_for_tier(schools, tier)
    row, err = icc.find_match(residence, tier, pool)
    if row is None:
        return None, err or "no match"
    sid = str(int(str(row.get("msid", "")).strip())).zfill(4)
    return sid, None


def find_denominator_column(ws) -> int:
    """1-based column index where row 7 contains 'Residents in Boundary'."""
    max_c = min(ws.max_column or 70, 80)
    for c in range(1, max_c + 1):
        v = ws.cell(7, c).value
        if v and "Residents in Boundary" in str(v):
            return c
    return 0


def forward_fill(ws, r: int, c_start: int, c_end: int) -> dict[int, str | None]:
    last: str | None = None
    out: dict[int, str | None] = {}
    for c in range(c_start, c_end + 1):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip() != "":
            last = str(v).strip()
        out[c] = last
    return out


def norm_header(x: object) -> str:
    if x is None:
        return ""
    s = str(x).replace("\r", " ").replace("\n", " ")
    return " ".join(s.split()).strip().casefold()


def cell_num(ws, r: int, c: int) -> float:
    v = ws.cell(r, c).value
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def classify_col(
    r6: str | None, r7: str | None, denom_c: int, c: int
) -> str:
    """district | choice | charter | skip"""
    if c == denom_c:
        return "skip"
    s7 = norm_header(r7)
    s6 = norm_header(r6) if r6 else ""
    if not s7:
        return "skip"
    if "non-geocoded" in s7 or "non-geocoded" in s6:
        return "skip"
    if "residents in boundary" in s7:
        return "skip"
    if "total number" in s7:
        return "skip"
    if "charter" in s6 or s7 == "charter schools":
        return "charter"
    if "choice" in s6:
        return "choice"
    if "brevard" in s6 and "district" in s6:
        return "district"
    return "skip"


def iter_residence_rows(ws, res_col: int = 3, res_col_fallback: int = 4):
    max_r = ws.max_row or 200
    for r in range(8, max_r + 5):
        raw = ws.cell(r, res_col).value
        if raw is None or str(raw).strip() == "":
            raw = ws.cell(r, res_col_fallback).value
        if raw is None or str(raw).strip() == "":
            continue
        s = str(raw).strip()
        if "Non-Geocoded" in s:
            continue
        if "Total Number" in s.replace("\n", " "):
            break
        if len(s) <= 2 and s.isalpha() and s.upper() == s:
            # e.g. footer "TO"
            continue
        yield r, s


def compute_workbook(path: Path, tier: str, schools: list[dict]) -> tuple[dict[str, dict], list[str]]:
    """
    Returns (by_msid aggregates, warnings).
    Each msid maps numerators + denominator sums (already summed if multiple rows).
    """
    warnings: list[str] = []
    agg: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "assignment": 0.0,
            "other_district": 0.0,
            "choice": 0.0,
            "charter": 0.0,
            "denominator": 0.0,
        }
    )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    denom_c = find_denominator_column(ws)
    if not denom_c:
        wb.close()
        warnings.append(f"No denominator column (Residents in Boundary) in {path.name}")
        return {}, warnings

    c_end = denom_c
    fill6 = forward_fill(ws, 6, 5, c_end)
    fill7 = forward_fill(ws, 7, 5, c_end)

    for r, residence in iter_residence_rows(ws):
        msid, err = match_residence_to_msid(residence, tier, schools)
        if not msid:
            warnings.append(f"{path.name} row {r} residence {residence!r}: {err}")
            continue

        den = cell_num(ws, r, denom_c)
        if den <= 0:
            warnings.append(
                f"{path.name} row {r} msid {msid} residence {residence!r}: denominator missing or zero (stored as 0)"
            )

        res_n = norm_header(residence)
        num_a = num_o = num_ch = num_char = 0.0

        for c in range(5, denom_c):
            cls = classify_col(fill6.get(c), fill7.get(c), denom_c, c)
            if cls == "skip":
                continue
            val = cell_num(ws, r, c)
            h7 = norm_header(fill7.get(c))
            if cls == "charter":
                num_char += val
            elif cls == "choice":
                num_ch += val
            elif cls == "district":
                if h7 == res_n:
                    num_a += val
                else:
                    num_o += val

        a = agg[msid]
        a["assignment"] += num_a
        a["other_district"] += num_o
        a["choice"] += num_ch
        a["charter"] += num_char
        a["denominator"] += den

    wb.close()
    return dict(agg), warnings


def int_count_str(x: float) -> str:
    return str(int(round(x)))


def rate_to_str(num: float, den: float) -> str:
    if den <= 0:
        return "0"
    x = num / den
    if x == 0:
        return "0"
    s = f"{x:.10f}".rstrip("0").rstrip(".")
    return s if s else "0"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--elem", default=DEFAULT_ELEM)
    ap.add_argument("--middle", default=DEFAULT_MID)
    ap.add_argument("--high", default=DEFAULT_HIGH)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    paths = {
        "elementary": Path(args.elem),
        "middle": Path(args.middle),
        "high": Path(args.high),
    }
    for tier, p in paths.items():
        if not p.is_file():
            print(f"Missing workbook ({tier}): {p}", file=sys.stderr)
            return 1

    if not CSV_PATH.is_file():
        print(f"Missing {CSV_PATH}", file=sys.stderr)
        return 1

    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        old_fields = reader.fieldnames or []
        schools = [dict(x) for x in reader]

    all_warn: list[str] = []
    combined: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "assignment": 0.0,
            "other_district": 0.0,
            "choice": 0.0,
            "charter": 0.0,
            "denominator": 0.0,
        }
    )

    for tier, p in paths.items():
        part, w = compute_workbook(p, tier, schools)
        all_warn.extend(w)
        for msid, vals in part.items():
            t = combined[msid]
            for k in vals:
                t[k] += vals[k]

    rate_slugs = [
        "assignment_capture_rate",
        "other_district_capture_rate",
        "choice_capture_rate",
        "charter_capture_rate",
    ]
    count_slugs = [
        "fromto_resident_denominator",
        "assignment_capture_students",
        "other_district_capture_students",
        "choice_capture_students",
        "charter_capture_students",
    ]
    new_slugs = rate_slugs + count_slugs

    def rebuild_fieldnames(h: list[str]) -> list[str]:
        drop = {"capture_rate", "charter_capture_rate", *new_slugs}
        h = [x for x in h if x and x not in drop]
        try:
            i = h.index("utilization_2025_26") + 1
        except ValueError:
            i = len(h)
        return h[:i] + new_slugs + h[i:]

    fieldnames = rebuild_fieldnames(list(old_fields))

    for row in schools:
        k = str(int(str(row.get("msid", "0")).strip())).zfill(4)
        vals = combined.get(k)
        if vals and vals["denominator"] > 0:
            den = vals["denominator"]
            row["assignment_capture_rate"] = rate_to_str(vals["assignment"], den)
            row["other_district_capture_rate"] = rate_to_str(
                vals["other_district"], den
            )
            row["choice_capture_rate"] = rate_to_str(vals["choice"], den)
            row["charter_capture_rate"] = rate_to_str(vals["charter"], den)
            row["fromto_resident_denominator"] = int_count_str(den)
            row["assignment_capture_students"] = int_count_str(vals["assignment"])
            row["other_district_capture_students"] = int_count_str(
                vals["other_district"]
            )
            row["choice_capture_students"] = int_count_str(vals["choice"])
            row["charter_capture_students"] = int_count_str(vals["charter"])
        else:
            if k in combined and combined[k]["denominator"] <= 0:
                all_warn.append(
                    f"MSID {k} {row.get('school_name', '')!r}: combined denominator <= 0; rates set to 0"
                )
            row["assignment_capture_rate"] = "0"
            row["other_district_capture_rate"] = "0"
            row["choice_capture_rate"] = "0"
            row["charter_capture_rate"] = "0"
            row["fromto_resident_denominator"] = "0"
            row["assignment_capture_students"] = "0"
            row["other_district_capture_students"] = "0"
            row["choice_capture_students"] = "0"
            row["charter_capture_students"] = "0"
        row.pop("capture_rate", None)
        # Keep charter_capture_rate: output column reuses the same header name as the legacy import.

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with REPORT_PATH.open("w", encoding="utf-8") as rf:
        rf.write("=== From-To capture import ===\n\n")
        for tier, p in paths.items():
            rf.write(f"{tier}: {p}\n")
        rf.write(f"\nMSIDs with From-To data: {len(combined)}\n\n")
        if all_warn:
            rf.write("Warnings:\n")
            for w in all_warn:
                rf.write(f"  - {w}\n")
        else:
            rf.write("Warnings: (none)\n")

    if args.dry_run:
        print("Dry run; not writing CSV. See warnings above.")
        for w in all_warn[:20]:
            print(w)
        return 0

    with CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(schools)

    print(f"Wrote {CSV_PATH} ({len(combined)} schools with source rows). Report: {REPORT_PATH}")
    if all_warn:
        print(f"{len(all_warn)} warning(s); see {REPORT_PATH}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
