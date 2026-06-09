# Reads data/raw/BPS Employee Count 5-6-26.xlsx (or pass path) → data/processed/bps_employee_count_by_msid.json
# Expected columns: MSID, LOCATION (or School Name), COUNT (or Count of employees).
# Run from repo root: py -3 scripts/export_bps_employee_count_from_xlsx.py
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError as e:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e


def _norm_msid(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v != v:
            return None
        i = int(v)
        return i if i == v or abs(v - i) < 1e-6 else None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _norm_count(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v != v:
            return None
        i = int(round(v))
        return i
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def _col_index_map(header_row: tuple) -> dict[str, int]:
    cols: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            cols[name.lower()] = i
    return cols


def read_workbook(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return {}

        cols = _col_index_map(header_row)

        def col(*names: str) -> Optional[int]:
            for nm in names:
                j = cols.get(nm.lower())
                if j is not None:
                    return j
            return None

        i_msid = col("msid")
        i_count = col("count", "count of employees", "employees", "employee count")
        if i_msid is None or i_count is None:
            raise SystemExit(
                f"Missing MSID or COUNT column in {path}. Headers: {list(header_row)}"
            )

        out: dict[str, int] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            mid = _norm_msid(row[i_msid] if i_msid < len(row) else None)
            cnt = _norm_count(row[i_count] if i_count < len(row) else None)
            if mid is None:
                continue
            if cnt is None:
                continue
            out[str(mid)] = cnt
        return out
    finally:
        wb.close()


def main() -> None:
    repo = Path(__file__).resolve().parents[1]
    default_src = repo / "data" / "raw" / "BPS Employee Count 5-6-26.xlsx"
    src = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else default_src
    if not src.is_file():
        raise SystemExit(f"Source not found: {src}")

    by_msid = read_workbook(src)
    out_path = repo / "data" / "processed" / "bps_employee_count_by_msid.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        src_rel = src.relative_to(repo)
    except ValueError:
        src_rel = src
    payload = {
        "sourceFile": str(src_rel),
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "notes": [
            "On-site BPS employee counts keyed by MSID (district export). Regenerate with scripts/export_bps_employee_count_from_xlsx.py.",
        ],
        "byMsid": by_msid,
    }
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {out_path} ({len(by_msid)} schools)")


if __name__ == "__main__":
    main()
