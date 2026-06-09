# Legacy one-time import: 251010ChartersFromAnalysis.xlsx -> charter_capture_rate (old schema).
# Dashboard capture KPIs are populated by scripts/import_fromto_capture_rates.py (From-To matrices).
# Sheet "PrintCharters", column D = short school name, column V = decimal fraction.
# PrintCharters row layout: rows 8–60 elementary, 63–74 middle, 75+ high (incl. jr/sr high).

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("Requires openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "data" / "school_master.csv"
DEFAULT_XLSX = r"P:\0109260\Planning\WorkingFiles\00_To Be Sorted\251010ChartersFromAnalysis.xlsx"
SHEET = "PrintCharters"
ROW_MID_START = 63
ROW_HIGH_START = 75

# Optional fixes when auto-match is ambiguous (norm(excel name) -> msid, no leading zeros)
MANUAL_EXCEL_NAME_TO_MSID: dict[str, str] = {}


def norm(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def word_tokens(s: str) -> list[str]:
    return re.findall(r"[A-Z0-9]+", s.upper(), flags=re.IGNORECASE)


def prefix_key(excel_name: str) -> str:
    n = norm(excel_name)
    if n.endswith(" ELEM"):
        n = n[:-5].rstrip() + " ELEMENTARY"
    return n


def tier_for_excel_row(r1: int) -> str | None:
    if r1 < ROW_MID_START:
        return "elementary"
    if r1 < ROW_HIGH_START:
        return "middle"
    return "high"  # includes jr_sr_high in CSV


def school_tier(s: dict) -> str | None:
    t = (s.get("school_level") or "").strip().lower()
    if t == "elementary":
        return "elementary"
    if t == "middle":
        return "middle"
    if t in ("high", "jr_sr_high"):
        return "high"
    return None


def pool_by_tier(schools: list[dict]) -> dict[str, list[dict]]:
    p: dict[str, list[dict]] = {"elementary": [], "middle": [], "high": []}
    for s in schools:
        st = school_tier(s)
        if st:
            p[st].append(s)
    return p


def find_match(
    d_s: str, tier: str, pool: list[dict]
) -> tuple[dict | None, str | None]:
    """
    Return (row, err). err set if no match or ambiguous.
    """
    nkey = norm(d_s)
    if nkey in MANUAL_EXCEL_NAME_TO_MSID:
        m = MANUAL_EXCEL_NAME_TO_MSID[nkey]
        for s in pool:
            sid = s.get("msid", "")
            sbase = str(int(sid)) if sid and str(sid).isdigit() else sid
            if sbase == m or str(sid).lstrip("0") == m.lstrip("0"):
                return s, None
    pfx = prefix_key(d_s)
    pfxn = norm(pfx)
    toks = word_tokens(pfxn)

    # 1) Strict prefix on school name
    by_pre: list[dict] = []
    for s in pool:
        sn = norm(s.get("school_name", ""))
        if pfxn and (sn == pfxn or sn.startswith(pfxn + " ")):
            by_pre.append(s)
    if len(by_pre) == 1:
        return by_pre[0], None
    if len(by_pre) > 1:
        by_pre2 = [s for s in by_pre if (s.get("appears_in_dropdown") or "").lower() == "yes"]
        if len(by_pre2) == 1:
            return by_pre2[0], None
        if pfxn == "COCOA" and tier == "high" and nkey == "COCOA":
            honly = [s for s in by_pre if "BEACH" not in norm(s.get("school_name", ""))]
            if len(honly) == 1:
                return honly[0], None
        return None, f"multiple prefix: {', '.join(x.get('school_name', '') for x in by_pre[:3])}"

    # 2) Single token: must match unique school in pool (e.g. last name of middle / high)
    if len(toks) == 1 and len(toks[0]) >= 2:
        t = toks[0]
        hit = [s for s in pool if t in word_tokens(s.get("school_name", ""))]
        if len(hit) == 1:
            return hit[0], None
        if len(hit) > 1:
            y = [s for s in hit if (s.get("appears_in_dropdown") or "").lower() == "yes"]
            if len(y) == 1:
                return y[0], None
            return None, f"token {t!r} -> {len(hit)} schools e.g. {hit[0].get('school_name')!r}"

    # 3) Two+ tokens: require all in school name, first word first
    if len(toks) >= 2:
        f0, f1 = toks[0], toks[1]
        hit = []
        for s in pool:
            name = s.get("school_name", "")
            u = name.upper()
            if f0 in u and f1 in u and u.find(f0) < u.find(f1):
                w = word_tokens(name)
                if f0 in w and f1 in w:
                    hit.append(s)
        if len(hit) == 1:
            return hit[0], None
        if len(hit) > 1:
            y = [s for s in hit if (s.get("appears_in_dropdown") or "").lower() == "yes"]
            if len(y) == 1:
                return y[0], None
            return None, f"pair {f0!r}/{f1!r} -> {len(hit)} schools"

    return None, "no match"


def float_to_cell(f: float) -> str:
    s = f"{f:.10f}".rstrip("0").rstrip(".")
    return s if s else "0"


def main() -> int:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
    if not xlsx.is_file():
        print(f"XLSX not found: {xlsx}", file=sys.stderr)
        return 1
    if not CSV_PATH.is_file():
        print(f"CSV not found: {CSV_PATH}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"Sheet {SHEET!r} not in {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb[SHEET]

    excel: list[tuple[int, str, float]] = []
    for r1 in range(1, (ws.max_row or 0) + 1):
        d = ws.cell(r1, 4).value
        v = ws.cell(r1, 22).value
        if d is None or d == "" or v is None or v == "":
            continue
        if isinstance(v, str) and "Percentage" in v:
            continue
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        d_s = str(d).strip()
        if not d_s:
            continue
        excel.append((r1, d_s, f))
    wb.close()

    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        h0 = r.fieldnames or []
        schools = [dict(x) for x in r]
    h0 = [x for x in h0 if x and x != "charter_capture_rate"]
    pbt = pool_by_tier(schools)

    by_msid: dict[str, str] = {}
    unmatched: list[str] = []
    matched: list[str] = []

    for r1, d_s, f in excel:
        tier = tier_for_excel_row(r1)
        if tier is None or tier not in pbt:
            unmatched.append(f"ROW {r1}\t{d_s!r} (no tier for row {r1})")
            continue
        row, err = find_match(d_s, tier, pbt[tier])
        if row is None:
            unmatched.append(f"ROW {r1}\t{d_s!r}\t(tier {tier}): {err or 'unmatched'}")
            continue
        k = (row.get("msid", "") or "").zfill(4)
        s = float_to_cell(f)
        if k in by_msid and by_msid[k] != s:
            print(f"Warning: MSID {k} duplicate in Excel, keeping last: {d_s!r}", file=sys.stderr)
        by_msid[k] = s
        matched.append(f"ROW {r1} {d_s!r} -> {k} {row.get('school_name')!r}")

    for row in schools:
        k = (row.get("msid", "") or "").zfill(4)
        if k in by_msid:
            row["charter_capture_rate"] = by_msid[k]
        else:
            row["charter_capture_rate"] = ""

    try:
        i_cap = h0.index("capture_rate")
    except ValueError:
        print("No capture_rate column in CSV", file=sys.stderr)
        return 1
    h_new = h0[: i_cap + 1] + ["charter_capture_rate"] + h0[i_cap + 1:]

    out = CSV_PATH
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=h_new, extrasaction="ignore")
        w.writeheader()
        w.writerows(schools)

    # Reports
    rpt = ROOT / "data" / "processed" / "charter_import_report.txt"
    rpt.parent.mkdir(parents=True, exist_ok=True)
    with rpt.open("w", encoding="utf-8") as rf:
        rf.write("=== Charter import from Excel (one-time) ===\n\n")
        rf.write(f"Source: {xlsx}\n\n")
        rf.write(f"Matched: {len(matched)} Excel rows into school_master charter_capture_rate.\n\n")
        rf.write(
            "P:\\ schools with no school_master match (unmatched short names; fix in Excel or add manual map in script):\n"
        )
        if unmatched:
            for u in unmatched:
                rf.write(f"  {u}\n")
        else:
            rf.write("  (none — all source rows joined.)\n")
        rf.write("\n--- Per-row join log ---\n")
        for line in sorted(matched, key=str.lower):
            rf.write(line + "\n")

    print("Wrote", out)
    print("Report", rpt)
    print("Matched", len(matched), "unmatched", len(unmatched))
    for u in unmatched:
        print(" UNMATCHED:", u)
    return 0 if not unmatched else 0


if __name__ == "__main__":
    raise SystemExit(main())
